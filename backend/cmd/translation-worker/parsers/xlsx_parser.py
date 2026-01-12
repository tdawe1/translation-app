"""
XLSX parser using openpyxl for Excel spreadsheets.

Extracts text from cells with coordinate and formatting preservation.
Handles multiple worksheets, formulas, and various data types.
"""

from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None
    Workbook = None
    get_column_letter = None

import sys
from pathlib import Path

worker_dir = Path(__file__).parent.parent
sys.path.insert(0, str(worker_dir))

from plugins import (
    ParsedDocument,
    Segment,
)


@dataclass
class CellInfo:
    """Information about an Excel cell."""

    coordinate: str
    row: int
    column: int
    value: Any
    is_formula: bool
    formula_text: Optional[str] = None
    data_type: str = "string"
    sheet_name: str = ""


@dataclass
class WorksheetInfo:
    """Information about an Excel worksheet."""

    name: str
    index: int
    cell_count: int
    has_data: bool


@dataclass
class FormattingInfo:
    """Text formatting information."""

    font_name: str = ""
    font_size: float = 0.0
    bold: bool = False
    italic: bool = False
    underline: bool = False
    color_rgb: Optional[Tuple[int, int, int]] = None


class XLSXParser:
    """Excel XLSX parser using openpyxl.

    Features:
    - Extract text from cells with coordinate context (A1, B2, etc.)
    - Preserve formatting (bold, italic, font size)
    - Handle multiple worksheets
    - Distinguish between values and formulas
    - Support various data types (text, numbers, dates)
    - Extract formulas as metadata (not translated by default)

    Note: openpyxl must be installed. Install with: pip install openpyxl
    """

    name = "xlsx_parser"
    version = "1.0.0"
    dependencies = ["openpyxl"]

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize XLSX parser.

        Args:
            config: Optional configuration dict

        Raises:
            RuntimeError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl is not installed. Install with: pip install openpyxl"
            )

        self.config = config or {}
        self._skip_formulas = self.config.get("skip_formulas", True)
        self._include_empty_sheets = self.config.get("include_empty_sheets", False)
        self._parse_all_sheets = self.config.get("parse_all_sheets", True)

    def initialize(self, config: Dict[str, Any]) -> None:
        """Initialize plugin with configuration.

        Args:
            config: Configuration dict
        """
        self.config.update(config)
        self._skip_formulas = self.config.get("skip_formulas", True)
        self._include_empty_sheets = self.config.get("include_empty_sheets", False)
        self._parse_all_sheets = self.config.get("parse_all_sheets", True)
        logger.info(f"[{self.name}] Initialized with config: {self.config}")

    def shutdown(self) -> None:
        """Clean up resources."""
        logger.debug(f"[{self.name}] Shutdown")

    def supported_extensions(self) -> List[str]:
        """Return list of supported file extensions.

        Returns:
            List of supported extensions
        """
        return [".xlsx", ".xlsm", ".xltx", ".xltm"]

    def parse(self, file_path: str) -> ParsedDocument:
        """Parse XLSX and extract translatable text segments.

        Args:
            file_path: Path to XLSX file

        Returns:
            ParsedDocument with segments

        Raises:
            FileNotFoundError: If file doesn't exist
            Exception: If XLSX cannot be parsed
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"XLSX file not found: {file_path}")

        wb = load_workbook(file_path, data_only=False)

        try:
            segments = []
            metadata = {
                "format": "xlsx",
                "sheet_count": len(wb.sheetnames),
            }

            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]

                worksheet_info = self._extract_worksheet_info(ws, sheet_idx)
                cells = self._extract_cells(ws, sheet_name)

                if self._include_empty_sheets or worksheet_info.has_data:
                    metadata[f"sheet_{sheet_name}_cells"] = len(cells)

                    for cell_info in cells:
                        segments.append(
                            self._create_cell_segment(cell_info, worksheet_info)
                        )

            return ParsedDocument(
                segments=segments,
                metadata=metadata,
                format="xlsx",
                source_path=file_path,
            )

        except Exception as e:
            logger.error(f"[{self.name}] Failed to parse XLSX: {e}")
            raise
        finally:
            pass

    def _extract_worksheet_info(self, ws, index: int) -> WorksheetInfo:
        """Extract information about a worksheet.

        Args:
            ws: openpyxl Worksheet object
            index: Worksheet index

        Returns:
            WorksheetInfo with sheet details
        """
        has_data = any(cell.value for row in ws.iter_rows() for cell in row)

        return WorksheetInfo(
            name=ws.title,
            index=index,
            cell_count=sum(1 for row in ws.iter_rows() for cell in row if cell.value),
            has_data=has_data,
        )

    def _extract_cells(self, ws, sheet_name: str) -> List[CellInfo]:
        """Extract all cells from a worksheet.

        Args:
            ws: openpyxl Worksheet object
            sheet_name: Name of the worksheet

        Returns:
            List of CellInfo objects
        """
        cells = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                is_formula = cell.data_type == "f"

                if is_formula and self._skip_formulas:
                    continue

                value = cell.value
                cell_type = self._get_cell_data_type(cell)

                cells.append(
                    CellInfo(
                        coordinate=cell.coordinate,
                        row=cell.row,
                        column=cell.column,
                        value=value,
                        is_formula=is_formula,
                        formula_text=cell.value if is_formula else None,
                        data_type=cell_type,
                        sheet_name=sheet_name,
                    )
                )

        return cells

    def _get_cell_data_type(self, cell) -> str:
        """Determine cell data type.

        Args:
            cell: openpyxl Cell object

        Returns:
            Data type string
        """
        from datetime import datetime, date

        if cell.is_date:
            return "date"
        elif isinstance(cell.value, (int, float)):
            return "numeric"
        elif isinstance(cell.value, str):
            return "string"
        elif isinstance(cell.value, bool):
            return "boolean"
        else:
            return "unknown"

    def _create_cell_segment(
        self, cell_info: CellInfo, worksheet_info: WorksheetInfo
    ) -> Segment:
        """Create a Segment from cell information.

        Args:
            cell_info: CellInfo object
            worksheet_info: WorksheetInfo object

        Returns:
            Segment with cell text and context
        """
        text = str(cell_info.value)

        metadata = {
            "is_formula": cell_info.is_formula,
            "formula_text": cell_info.formula_text,
            "data_type": cell_info.data_type,
        }

        context = {
            "type": "cell",
            "sheet_name": cell_info.sheet_name,
            "coordinate": cell_info.coordinate,
            "row": cell_info.row,
            "column": cell_info.column,
        }

        return Segment(
            id=f"{cell_info.sheet_name}!{cell_info.coordinate}",
            text=text,
            context=context,
            metadata=metadata,
        )

    def render(
        self, doc: ParsedDocument, output_path: str, template_path: Optional[str] = None
    ) -> None:
        """Render translated XLSX with original layout.

        Args:
            doc: ParsedDocument with translated segments
            output_path: Where to save output XLSX
            template_path: Optional original XLSX as template

        Raises:
            Exception: If rendering fails
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        if template_path:
            try:
                wb = load_workbook(template_path)
            except Exception as e:
                logger.warning(f"[{self.name}] Failed to load template: {e}")
                wb = Workbook()
        else:
            wb = Workbook()

        try:
            for segment in doc.segments:
                sheet_name = segment.context.get("sheet_name", "Sheet")
                coordinate = segment.context.get("coordinate")

                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)

                ws = wb[sheet_name]

                if coordinate:
                    ws[coordinate] = segment.text

            wb.save(output_path)
            logger.info(f"[{self.name}] Rendered XLSX to: {output_path}")

        except Exception as e:
            logger.error(f"[{self.name}] Failed to render XLSX: {e}")
            raise

    def get_worksheet_count(self, file_path: str) -> int:
        """Get number of worksheets in a workbook.

        Args:
            file_path: Path to XLSX file

        Returns:
            Number of worksheets
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed")

        wb = load_workbook(file_path, read_only=True)
        return len(wb.sheetnames)


def create_xlsx_parser(config: Optional[Dict[str, Any]] = None) -> XLSXParser:
    """Factory function to create an XLSX parser.

    Args:
        config: Optional configuration dict

    Returns:
        XLSXParser instance

    Raises:
        RuntimeError: If openpyxl is not installed
    """
    return XLSXParser(config)
