# tests/test_parsers/test_xlsx_parser.py
"""
Unit tests for XLSX parser functionality.

Tests openpyxl-based Excel parsing, extraction, and rendering.
Tests gracefully skip when openpyxl is not installed.
"""

import pytest
import sys
import tempfile
from pathlib import Path
from datetime import datetime

# Add worker directory to path for imports
worker_dir = Path(__file__).parent.parent.parent
sys.path.insert(0, str(worker_dir))

# Import parser module
try:
    from parsers.xlsx_parser import (
        XLSXParser,
        CellInfo,
        WorksheetInfo,
        create_xlsx_parser,
        OPENPYXL_AVAILABLE,
    )

    OPENPYXL_INSTALLED = OPENPYXL_AVAILABLE
except ImportError:
    OPENPYXL_INSTALLED = False
    XLSXParser = None
    CellInfo = None
    WorksheetInfo = None
    create_xlsx_parser = None


@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="openpyxl not installed")
class TestCellInfo:
    """Test CellInfo dataclass."""

    def test_create_cell_info(self):
        """Should create CellInfo with correct attributes."""
        cell = CellInfo(
            coordinate="A1",
            row=1,
            column=1,
            value="Hello",
            is_formula=False,
            formula_text=None,
            data_type="string",
            sheet_name="Sheet1",
        )

        assert cell.coordinate == "A1"
        assert cell.row == 1
        assert cell.column == 1
        assert cell.value == "Hello"
        assert cell.is_formula is False
        assert cell.formula_text is None
        assert cell.data_type == "string"
        assert cell.sheet_name == "Sheet1"

    def test_cell_info_with_formula(self):
        """Should create CellInfo with formula."""
        cell = CellInfo(
            coordinate="B2",
            row=2,
            column=2,
            value=42,
            is_formula=True,
            formula_text="=SUM(A1:A2)",
            data_type="numeric",
            sheet_name="Sheet1",
        )

        assert cell.is_formula is True
        assert cell.formula_text == "=SUM(A1:A2)"
        assert cell.value == 42


@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="openpyxl not installed")
class TestWorksheetInfo:
    """Test WorksheetInfo dataclass."""

    def test_create_worksheet_info(self):
        """Should create WorksheetInfo with correct attributes."""
        ws = WorksheetInfo(name="Sheet1", index=0, cell_count=5, has_data=True)

        assert ws.name == "Sheet1"
        assert ws.index == 0
        assert ws.cell_count == 5
        assert ws.has_data is True


@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="openpyxl not installed")
class TestXLSXParser:
    """Test XLSXParser functionality."""

    def test_plugin_attributes(self):
        """Should have required plugin attributes."""
        parser = XLSXParser()

        assert hasattr(parser, "name")
        assert hasattr(parser, "version")
        assert hasattr(parser, "dependencies")
        assert parser.name == "xlsx_parser"
        assert parser.version == "1.0.0"
        assert "openpyxl" in parser.dependencies

    def test_supported_extensions(self):
        """Should return XLSX and related extensions."""
        parser = XLSXParser()

        extensions = parser.supported_extensions()
        assert ".xlsx" in extensions
        assert ".xlsm" in extensions
        assert len(extensions) >= 2

    def test_initialization_default_config(self):
        """Should initialize with default config."""
        parser = XLSXParser()

        assert parser._skip_formulas is True
        assert parser._include_empty_sheets is False
        assert parser._parse_all_sheets is True

    def test_initialization_custom_config(self):
        """Should accept custom configuration."""
        parser = XLSXParser(
            config={
                "skip_formulas": False,
                "include_empty_sheets": True,
                "parse_all_sheets": False,
            }
        )

        assert parser._skip_formulas is False
        assert parser._include_empty_sheets is True
        assert parser._parse_all_sheets is False

    def test_initialize_method(self):
        """Should update config via initialize method."""
        parser = XLSXParser()
        parser.initialize({"skip_formulas": False, "parse_all_sheets": False})

        assert parser._skip_formulas is False
        assert parser._parse_all_sheets is False

    def test_shutdown(self):
        """Should shutdown without errors."""
        parser = XLSXParser()
        parser.shutdown()  # Should not raise

    def test_parse_nonexistent_file(self):
        """Should raise FileNotFoundError for missing file."""
        parser = XLSXParser()

        with pytest.raises(FileNotFoundError):
            parser.parse("/nonexistent/file.xlsx")

    def test_create_xlsx_parser_factory(self):
        """Should create parser via factory function."""
        parser = create_xlsx_parser()

        assert isinstance(parser, XLSXParser)

    def test_create_xlsx_parser_with_config(self):
        """Should create parser with config via factory."""
        parser = create_xlsx_parser({"skip_formulas": False})

        assert parser._skip_formulas is False


@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="openpyxl not installed")
class TestXLSXParserWithSampleFile:
    """Test parser with actual XLSX file."""

    def test_parse_simple_xlsx(self):
        """Should parse a simple XLSX file."""
        from openpyxl import Workbook

        parser = XLSXParser()

        # Create a simple test XLSX
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Hello World"
            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                assert parsed.format == "xlsx"
                assert len(parsed.segments) >= 1
                assert parsed.segments[0].text == "Hello World"
                assert parsed.segments[0].context.get("coordinate") == "A1"
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_parse_multiple_cells(self):
        """Should parse multiple cells."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = "First"
            ws["B1"] = "Second"
            ws["C1"] = "Third"

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                assert len(parsed.segments) >= 3
                texts = [s.text for s in parsed.segments]
                assert "First" in texts
                assert "Second" in texts
                assert "Third" in texts
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_cell_reference_in_metadata(self):
        """Should include cell reference in context."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = "Cell A1"
            ws["B2"] = "Cell B2"
            ws["C3"] = "Cell C3"

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Check that cell references are in context
                cell_a1 = next(
                    (s for s in parsed.segments if s.context.get("coordinate") == "A1"),
                    None,
                )
                cell_b2 = next(
                    (s for s in parsed.segments if s.context.get("coordinate") == "B2"),
                    None,
                )
                cell_c3 = next(
                    (s for s in parsed.segments if s.context.get("coordinate") == "C3"),
                    None,
                )

                assert cell_a1 is not None
                assert cell_b2 is not None
                assert cell_c3 is not None

                assert cell_a1.context.get("row") == 1
                assert cell_a1.context.get("column") == 1
                assert cell_a1.context.get("sheet_name") == "Sheet"
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_parse_with_formula_skipped(self):
        """Should skip formulas by default."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = "Value"
            ws["B1"] = 10
            ws["C1"] = 20
            ws["D1"] = "=SUM(B1:C1)"  # Formula

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Formula should not be in segments by default
                formula_segment = next(
                    (s for s in parsed.segments if s.context.get("coordinate") == "D1"),
                    None,
                )
                assert (
                    formula_segment is None
                    or formula_segment.metadata.get("is_formula") is False
                )
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_parse_with_formula_included(self):
        """Should include formulas when configured."""
        from openpyxl import Workbook

        parser = XLSXParser(config={"skip_formulas": False})

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = "Value"
            ws["B1"] = 10
            ws["C1"] = 20
            ws["D1"] = "=SUM(B1:C1)"  # Formula

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Formula should be in segments
                formula_segment = next(
                    (s for s in parsed.segments if s.context.get("coordinate") == "D1"),
                    None,
                )
                assert formula_segment is not None
                assert formula_segment.metadata.get("is_formula") is True
                assert formula_segment.metadata.get("formula_text") == "=SUM(B1:C1)"
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_parse_multiple_sheets(self):
        """Should parse multiple worksheets."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()

            # First sheet
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws1["A1"] = "Data from Sheet1"

            # Second sheet
            ws2 = wb.create_sheet("Sheet2")
            ws2["A1"] = "Data from Sheet2"

            # Third sheet
            ws3 = wb.create_sheet("Sheet3")
            ws3["A1"] = "Data from Sheet3"

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                assert parsed.metadata.get("sheet_count") == 3

                # Check data from each sheet
                sheet1_data = [
                    s
                    for s in parsed.segments
                    if s.context.get("sheet_name") == "Sheet1"
                ]
                sheet2_data = [
                    s
                    for s in parsed.segments
                    if s.context.get("sheet_name") == "Sheet2"
                ]
                sheet3_data = [
                    s
                    for s in parsed.segments
                    if s.context.get("sheet_name") == "Sheet3"
                ]

                assert len(sheet1_data) > 0
                assert len(sheet2_data) > 0
                assert len(sheet3_data) > 0

                assert "Data from Sheet1" in sheet1_data[0].text
                assert "Data from Sheet2" in sheet2_data[0].text
                assert "Data from Sheet3" in sheet3_data[0].text
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_handle_numeric_cells(self):
        """Should handle numeric cell values."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = 42
            ws["B1"] = 3.14159
            ws["C1"] = -100

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Numeric cells should be converted to strings
                assert len(parsed.segments) >= 3
                assert "42" in [s.text for s in parsed.segments]
                assert "3.14159" in [s.text for s in parsed.segments]
                assert "-100" in [s.text for s in parsed.segments]
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_handle_date_cells(self):
        """Should handle date cell values."""
        from openpyxl import Workbook
        from datetime import date

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = date(2024, 1, 1)
            ws["B1"] = date(2024, 12, 31)

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Date cells should be converted to strings
                assert len(parsed.segments) >= 2
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_skip_empty_cells(self):
        """Should skip empty cells."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = "Data"
            ws["B1"] = None  # Empty
            ws["C1"] = ""  # Empty string

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Should only have non-empty cells
                assert len(parsed.segments) == 1
                assert parsed.segments[0].text == "Data"
            finally:
                Path(tmp.name).unlink(missing_ok=True)


@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="openpyxl not installed")
class TestXLSXParserRendering:
    """Test XLSX rendering functionality."""

    def test_render_with_template(self):
        """Should render XLSX using template."""
        from openpyxl import Workbook
        from plugins import ParsedDocument, Segment

        parser = XLSXParser()

        # Create source XLSX
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as source_tmp:
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False
            ) as output_tmp:
                wb = Workbook()
                ws = wb.active

                ws["A1"] = "Original"
                ws["B1"] = "Content"

                wb.save(source_tmp.name)

                try:
                    # Create parsed document with translation
                    parsed = ParsedDocument(
                        segments=[
                            Segment(
                                id="Sheet!A1",
                                text="Translated",
                                context={
                                    "type": "cell",
                                    "sheet_name": "Sheet",
                                    "coordinate": "A1",
                                    "row": 1,
                                    "column": 1,
                                },
                                metadata={},
                            ),
                            Segment(
                                id="Sheet!B1",
                                text="Translated Content",
                                context={
                                    "type": "cell",
                                    "sheet_name": "Sheet",
                                    "coordinate": "B1",
                                    "row": 1,
                                    "column": 2,
                                },
                                metadata={},
                            ),
                        ],
                        metadata={"format": "xlsx", "sheet_count": 1},
                        format="xlsx",
                        source_path=source_tmp.name,
                    )

                    # Render with template
                    parser.render(
                        doc=parsed,
                        output_path=output_tmp.name,
                        template_path=source_tmp.name,
                    )

                    # Verify output file exists and has content
                    assert Path(output_tmp.name).exists()
                    assert Path(output_tmp.name).stat().st_size > 0

                    # Verify we can open rendered file
                    from openpyxl import load_workbook

                    out_wb = load_workbook(output_tmp.name)
                    assert len(out_wb.sheetnames) >= 1

                    # Check translated values
                    out_ws = out_wb.active
                    assert out_ws["A1"].value == "Translated"
                    assert out_ws["B1"].value == "Translated Content"

                finally:
                    Path(source_tmp.name).unlink(missing_ok=True)
                    Path(output_tmp.name).unlink(missing_ok=True)

    def test_render_without_template(self):
        """Should render XLSX without template (new document)."""
        from openpyxl import Workbook
        from plugins import ParsedDocument, Segment

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_tmp:
            try:
                parsed = ParsedDocument(
                    segments=[
                        Segment(
                            id="Sheet!A1",
                            text="Test content",
                            context={
                                "type": "cell",
                                "sheet_name": "Sheet",
                                "coordinate": "A1",
                                "row": 1,
                                "column": 1,
                            },
                            metadata={},
                        )
                    ],
                    metadata={"format": "xlsx"},
                    format="xlsx",
                )

                parser.render(doc=parsed, output_path=output_tmp.name)

                assert Path(output_tmp.name).exists()
                assert Path(output_tmp.name).stat().st_size > 0

            finally:
                Path(output_tmp.name).unlink(missing_ok=True)

    def test_render_multiple_sheets(self):
        """Should render translated content across multiple sheets."""
        from openpyxl import Workbook
        from plugins import ParsedDocument, Segment

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as source_tmp:
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False
            ) as output_tmp:
                wb = Workbook()

                ws1 = wb.active
                ws1.title = "Sheet1"
                ws1["A1"] = "Original 1"

                ws2 = wb.create_sheet("Sheet2")
                ws2["A1"] = "Original 2"

                wb.save(source_tmp.name)

                try:
                    parsed = ParsedDocument(
                        segments=[
                            Segment(
                                id="Sheet1!A1",
                                text="Translated 1",
                                context={
                                    "type": "cell",
                                    "sheet_name": "Sheet1",
                                    "coordinate": "A1",
                                    "row": 1,
                                    "column": 1,
                                },
                                metadata={},
                            ),
                            Segment(
                                id="Sheet2!A1",
                                text="Translated 2",
                                context={
                                    "type": "cell",
                                    "sheet_name": "Sheet2",
                                    "coordinate": "A1",
                                    "row": 1,
                                    "column": 1,
                                },
                                metadata={},
                            ),
                        ],
                        metadata={"format": "xlsx", "sheet_count": 2},
                        format="xlsx",
                        source_path=source_tmp.name,
                    )

                    parser.render(
                        doc=parsed,
                        output_path=output_tmp.name,
                        template_path=source_tmp.name,
                    )

                    # Verify output
                    from openpyxl import load_workbook

                    out_wb = load_workbook(output_tmp.name)

                    assert "Sheet1" in out_wb.sheetnames
                    assert "Sheet2" in out_wb.sheetnames

                    assert out_wb["Sheet1"]["A1"].value == "Translated 1"
                    assert out_wb["Sheet2"]["A1"].value == "Translated 2"

                finally:
                    Path(source_tmp.name).unlink(missing_ok=True)
                    Path(output_tmp.name).unlink(missing_ok=True)


@pytest.mark.skipif(not OPENPYXL_INSTALLED, reason="openpyxl not installed")
class TestXLSXParserEdgeCases:
    """Test edge cases and special scenarios."""

    def test_empty_workbook(self):
        """Should handle empty workbook."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            # Empty sheet with no data

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Empty workbook should have no segments
                assert parsed.format == "xlsx"
                assert len(parsed.segments) == 0
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_empty_sheets(self):
        """Should skip empty sheets by default."""
        from openpyxl import Workbook

        parser = XLSXParser(config={"include_empty_sheets": False})

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()

            # Sheet with data
            ws1 = wb.active
            ws1.title = "DataSheet"
            ws1["A1"] = "Data"

            # Empty sheet
            ws2 = wb.create_sheet("EmptySheet")

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                # Should only have data from non-empty sheet
                assert len(parsed.segments) == 1
                assert parsed.segments[0].text == "Data"
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_unicode_content(self):
        """Should handle unicode characters."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            ws["A1"] = "日本語"
            ws["B1"] = "Русский"
            ws["C1"] = "🎉 Emoji"

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                text = " ".join(s.text for s in parsed.segments)
                assert "日本語" in text
                assert "Русский" in text
                assert "🎉" in text
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_long_text_in_cell(self):
        """Should handle long text in cells."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            ws = wb.active

            long_text = "Word " * 1000
            ws["A1"] = long_text

            wb.save(tmp.name)

            try:
                parsed = parser.parse(tmp.name)

                assert len(parsed.segments) == 1
                assert len(parsed.segments[0].text) > 4000
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    def test_get_worksheet_count(self):
        """Should return correct worksheet count."""
        from openpyxl import Workbook

        parser = XLSXParser()

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = Workbook()
            wb.active.title = "Sheet1"
            wb.create_sheet("Sheet2")
            wb.create_sheet("Sheet3")
            wb.save(tmp.name)

            try:
                count = parser.get_worksheet_count(tmp.name)
                assert count == 3
            finally:
                Path(tmp.name).unlink(missing_ok=True)


class TestXLSXNotInstalled:
    """Test behavior when openpyxl is not installed."""

    def test_parser_raises_when_not_installed(self):
        """Should raise RuntimeError when openpyxl unavailable."""
        if OPENPYXL_INSTALLED:
            pytest.skip("openpyxl is installed")

        with pytest.raises(RuntimeError, match="openpyxl is not installed"):
            XLSXParser()

    def test_factory_raises_when_not_installed(self):
        """Should raise RuntimeError from factory when openpyxl unavailable."""
        if OPENPYXL_INSTALLED:
            pytest.skip("openpyxl is installed")

        with pytest.raises(RuntimeError, match="openpyxl is not installed"):
            create_xlsx_parser()
